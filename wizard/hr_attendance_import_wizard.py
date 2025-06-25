# models/hr_attendance_import_wizard.py

import base64
import io
import logging
from datetime import datetime, timedelta, time

import pytz
from openpyxl import load_workbook
from odoo import models, fields

_logger = logging.getLogger(__name__)


class HrAttendanceImport(models.TransientModel):
    _name = "hr.attendance.import"
    _description = "Importar Asistencias desde Excel"

    file_data = fields.Binary("Archivo Excel", required=True)
    file_name = fields.Char("Nombre de archivo")
    allow_partial_attendance = fields.Boolean(
        "Permitir asistencias parciales",
        default=False,
        help="Crear asistencias solo con entrada cuando falte la salida"
    )

    def _get_work_schedule_for_date(self, employee, target_date):
        """
        Obtiene el horario de trabajo del empleado para una fecha específica
        """
        calendar = employee.resource_calendar_id
        if not calendar:
            return None

        # Obtener el día de la semana (0=Lunes, 6=Domingo)
        weekday = target_date.weekday()

        # Buscar la línea de horario correspondiente al día
        attendance_lines = calendar.attendance_ids.filtered(
            lambda line: int(line.dayofweek) == weekday
        )

        if not attendance_lines:
            return None

        # Para turnos nocturnos, puede haber múltiples líneas
        # Ordenar por hora de inicio
        attendance_lines = attendance_lines.sorted('hour_from')

        schedule = []
        for line in attendance_lines:
            schedule.append({
                'hour_from': line.hour_from,
                'hour_to': line.hour_to,
                'name': line.name or f"{line.dayofweek} - {line.display_name}"
            })

        return schedule

    def _group_marks_by_work_shift(self, times_sorted, schedule_info):
        """
        Agrupa las marcas por turno de trabajo, considerando turnos nocturnos
        que abarcan dos días calendario
        """
        if not schedule_info:
            # Sin información de horario, agrupar por día simple
            daily_marks = {}
            for t in times_sorted:
                work_date = t.date()
                if work_date not in daily_marks:
                    daily_marks[work_date] = []
                daily_marks[work_date].append(t)
            return daily_marks

        calendar = schedule_info['calendar']
        is_night_shift = calendar.nocturna

        if not is_night_shift:
            # Turno diurno: agrupar por día simple
            _logger.info("  Turno diurno: Agrupando por día calendario")
            daily_marks = {}
            for t in times_sorted:
                work_date = t.date()
                if work_date not in daily_marks:
                    daily_marks[work_date] = []
                daily_marks[work_date].append(t)
            return daily_marks

        # Turno nocturno: agrupar inteligentemente
        _logger.info("  Turno nocturno: Agrupando por turnos que cruzan medianoche")

        shift_groups = {}

        for mark in times_sorted:
            mark_hour = mark.hour + (mark.minute / 60.0)

            # Determinar a qué turno pertenece esta marca
            if mark_hour >= 17.0:  # Marca de tarde/noche (17:00-23:59)
                # Esta es una entrada, el turno se identifica por este día
                shift_date = mark.date()
                _logger.info("    Marca de entrada nocturna %s -> Turno del %s", mark,
                             shift_date)
            elif mark_hour <= 10.0:  # Marca de madrugada (00:00-10:00)
                # Esta es una salida, el turno se identifica por el día anterior
                shift_date = mark.date() - timedelta(days=1)
                _logger.info(
                    "    Marca de salida nocturna %s -> Turno del %s (día anterior)",
                    mark, shift_date)
            else:  # Marca de día (10:00-17:00) - caso raro en turno nocturno
                # Probablemente una salida tardía, asociar al día anterior
                shift_date = mark.date() - timedelta(days=1)
                _logger.info(
                    "    Marca de día %s -> Turno del %s (día anterior, salida tardía)",
                    mark, shift_date)

            # Agregar marca al grupo del turno
            if shift_date not in shift_groups:
                shift_groups[shift_date] = []
            shift_groups[shift_date].append(mark)

        _logger.info("  Turnos nocturnos agrupados:")
        for shift_date, marks in shift_groups.items():
            _logger.info("    Turno %s: %d marcas -> %s", shift_date, len(marks), marks)

        return shift_groups

    def _analyze_employee_schedule(self, employee):
        """
        Analiza el horario del empleado y determina la estrategia de procesamiento
        """
        calendar = employee.resource_calendar_id
        if not calendar:
            return None

        # Obtener información del horario
        schedule_info = {
            'employee': employee,
            'calendar': calendar,
            'is_night_shift': calendar.nocturna,
            'daily_schedules': {},
            'expected_entry_hours': [],
            'expected_exit_hours': []
        }

        # Analizar horario por cada día de la semana
        for weekday in range(7):  # 0=Lunes, 6=Domingo
            day_schedule = self._get_work_schedule_for_date(employee, datetime(2025, 6,
                                                                               2 + weekday).date())
            if day_schedule:
                schedule_info['daily_schedules'][weekday] = day_schedule

                # Recopilar horas típicas de entrada y salida
                for block in day_schedule:
                    schedule_info['expected_entry_hours'].append(block['hour_from'])
                    schedule_info['expected_exit_hours'].append(block['hour_to'])

        # Determinar rangos típicos
        if schedule_info['expected_entry_hours']:
            schedule_info['typical_entry_start'] = min(
                schedule_info['expected_entry_hours'])
            schedule_info['typical_entry_end'] = max(
                schedule_info['expected_entry_hours'])

        if schedule_info['expected_exit_hours']:
            schedule_info['typical_exit_start'] = min(
                schedule_info['expected_exit_hours'])
            schedule_info['typical_exit_end'] = max(schedule_info['expected_exit_hours'])

        _logger.info("Análisis de horario para %s: %s", employee.name, {
            'nocturno': schedule_info['is_night_shift'],
            'entrada_típica': f"{schedule_info.get('typical_entry_start', 0):.1f}h",
            'salida_típica': f"{schedule_info.get('typical_exit_end', 0):.1f}h"
        })

        return schedule_info

    def _classify_marks_by_schedule(self, marks, schedule_info):
        """
        Clasifica las marcas según el horario del empleado
        """
        if not schedule_info:
            return {'entries': marks[:1] if marks else [],
                    'exits': marks[1:] if len(marks) > 1 else [], 'unknown': []}

        classified = {
            'entries': [],
            'exits': [],
            'unknown': []
        }

        is_night_shift = schedule_info['is_night_shift']

        _logger.info("    Clasificando marcas para turno %s",
                     "nocturno" if is_night_shift else "diurno")

        for mark in marks:
            mark_hour = mark.hour + (mark.minute / 60.0)
            _logger.info("    Evaluando marca %s (hora: %.2f)", mark, mark_hour)

            if is_night_shift:
                # Turno nocturno: típicamente 16:00-06:00 o 18:00-06:00
                # Entradas: entre 16:00-23:59 (tarde/noche del día actual)
                # Salidas: entre 00:00-10:00 (madrugada del día siguiente)

                if mark_hour >= 16.0:  # 4PM en adelante = entrada
                    _logger.info("    Marca de tarde/noche (>=16h) -> ENTRADA")
                    classified['entries'].append(mark)
                elif mark_hour <= 07.0:  # Madrugada hasta 07AM = salida
                    _logger.info("    Marca de madrugada (<=07h) -> SALIDA")
                    classified['exits'].append(mark)
                else:  # Entre 07AM y 4PM = hora rara, probablemente salida
                    _logger.info("    Marca de día (07-16h) -> SALIDA")
                    classified['exits'].append(mark)
            else:
                # Turno diurno: entrada en la mañana, salida en la tarde
                if mark_hour <= 10.0:  # Mañana = entrada
                    _logger.info("    Marca de mañana (<=10h) -> ENTRADA")
                    classified['entries'].append(mark)
                elif mark_hour > 10.0:  # Tarde = salida
                    _logger.info("    Marca de tarde (>10h) -> SALIDA")
                    classified['exits'].append(mark)
                else:
                    _logger.info("Marca no clasificable -> DESCONOCIDA")
                    classified['unknown'].append(mark)

        _logger.info(
            "  Resultado clasificación: Entradas=%s, Salidas=%s, Desconocidas=%s",
            len(classified['entries']), len(classified['exits']),
            len(classified['unknown']))

        return classified

    def _find_existing_partial_attendance(self, employee, work_date):
        """
        Busca asistencias parciales existentes para completar
        """
        day_start = datetime.combine(work_date, datetime.min.time())
        day_end = day_start + timedelta(days=1)

        existing_partial = self.env["hr.attendance"].search([
            ("employee_id", "=", employee.id),
            ("check_in", ">=", day_start.strftime("%Y-%m-%d %H:%M:%S")),
            ("check_in", "<", day_end.strftime("%Y-%m-%d %H:%M:%S")),
            ("is_partial", "=", True),
            ("partial_type", "=", "entry_only")
        ], limit=1)

        return existing_partial

    def _check_duplicate_attendance(self, employee, check_datetime):
        """
        Verifica si ya existe una asistencia completa para el mismo empleado y día
        """
        _logger.info("    Verificando duplicados para %s en %s", employee.name,
                     check_datetime)

        # Obtener el inicio y fin del día en timezone local
        user_tz = self.env.user.tz or 'UTC'
        local_tz = pytz.timezone(user_tz)
        utc_tz = pytz.utc

        # Convertir a timezone local para obtener el día correcto
        if check_datetime.tzinfo is None:
            local_dt = local_tz.localize(check_datetime)
        else:
            local_dt = check_datetime.astimezone(local_tz)

        day_start = local_dt.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)

        # Convertir de vuelta a UTC para la búsqueda
        day_start_utc = day_start.astimezone(utc_tz)
        day_end_utc = day_end.astimezone(utc_tz)

        _logger.info("    Buscando duplicados entre %s y %s", day_start_utc, day_end_utc)

        # Buscar asistencias completas (no parciales)
        existing = self.env["hr.attendance"].search([
            ("employee_id", "=", employee.id),
            ("check_in", ">=", day_start_utc.strftime("%Y-%m-%d %H:%M:%S")),
            ("check_in", "<", day_end_utc.strftime("%Y-%m-%d %H:%M:%S")),
            ("is_partial", "=", False)  # Solo asistencias completas
        ])

        _logger.info("    Registros completos existentes: %d", len(existing))
        if existing:
            for rec in existing:
                _logger.info("    - ID %d: %s - %s", rec.id, rec.check_in, rec.check_out)

        return len(existing) > 0

    def _calculate_theoretical_entry_time(self, employee, work_date, exit_time,
                                          schedule_info):
        """
        Calcula la hora de entrada teórica basada en el horario de trabajo
        cuando solo se tiene la marca de salida
        """
        _logger.info("    Calculando entrada teórica para %s en fecha %s", employee.name,
                     work_date)
        _logger.info("    Hora de salida disponible: %s", exit_time)

        if not schedule_info:
            _logger.warning("    No hay información de horario disponible")
            return None

        calendar = schedule_info['calendar']
        user_tz = self.env.user.tz or 'UTC'
        local_tz = pytz.timezone(user_tz)

        # Convertir exit_time a timezone local si es necesario
        if exit_time.tzinfo is None:
            local_exit = local_tz.localize(exit_time)
        else:
            local_exit = exit_time.astimezone(local_tz)

        target_date = work_date

        # Para turnos nocturnos, ajustar la fecha objetivo si es necesario
        if calendar.nocturna:
            _logger.info("    Turno nocturno detectado")
            # Si la salida es en la madrugada (antes de las 12 PM),
            # el turno empezó el día anterior
            if local_exit.hour < 12:
                target_date = work_date - timedelta(days=1)
                _logger.info(
                    "    Salida de madrugada, buscando horario del día anterior: %s",
                    target_date)

        # Obtener horario de trabajo para la fecha objetivo
        schedule = self._get_work_schedule_for_date(employee, target_date)
        if not schedule:
            _logger.warning("    No se encontró horario para fecha %s", target_date)
            return None

        _logger.info("    Horarios encontrados: %s", schedule)

        # Buscar el horario más apropiado según la hora de salida
        best_match = None
        min_diff = float('inf')

        for i, sched in enumerate(schedule):
            # Convertir hora de inicio del horario a datetime
            hour_from_decimal = sched['hour_from']
            hours = int(hour_from_decimal)
            minutes = int((hour_from_decimal - hours) * 60)

            # Crear datetime para la hora de inicio del turno
            theoretical_entry = local_exit.replace(hour=hours, minute=minutes, second=0,
                                                   microsecond=0)

            # Para turnos nocturnos, ajustar la fecha de entrada si es necesario
            if calendar.nocturna:
                if hours >= 16:  # Turno que inicia en la tarde (desde 16:00)
                    if local_exit.hour < 12:  # Si la salida es en la madrugada
                        # La entrada fue el día anterior
                        theoretical_entry = theoretical_entry - timedelta(days=1)
                        _logger.info(
                            "    Ajustado para turno nocturno: entrada el día anterior")

            # Calcular si esta combinación de horarios tiene sentido
            # La entrada teórica debe ser antes de la salida real
            if theoretical_entry < local_exit:
                time_diff = (local_exit - theoretical_entry).total_seconds()
                _logger.info(
                    "    Horario %d: Entrada teórica %s -> Salida real %s (Duración: %.1f horas)",
                    i + 1, theoretical_entry, local_exit, time_diff / 3600)

                # Preferir horarios que resulten en jornadas laborales razonables (4-12 horas)
                if 4 * 3600 <= time_diff <= 12 * 3600:  # Entre 4 y 12 horas
                    if time_diff < min_diff:
                        min_diff = time_diff
                        best_match = theoretical_entry
                        _logger.info("    ✓ Nuevo mejor match (jornada razonable): %s",
                                     best_match)
                else:
                    _logger.info("    Jornada fuera de rango razonable (%.1f horas)",
                                 time_diff / 3600)
            else:
                _logger.info(
                    "    Horario %d descartado: entrada teórica posterior a salida",
                    i + 1)

        if best_match:
            _logger.info("    ✅ Entrada teórica calculada: %s", best_match)
            return best_match.replace(tzinfo=None)
        else:
            _logger.warning("    ❌ No se pudo calcular entrada teórica válida")
            return None

    def _adjust_check_in_time(self, employee, original_datetime):
        """
        Ajusta la hora de entrada según el horario de trabajo del empleado
        SOLO si llega ANTES de su horario (no después)
        """
        _logger.info("    Ajustando hora de entrada para %s", employee.name)
        _logger.info("    Hora original: %s", original_datetime)

        user_tz = self.env.user.tz or 'UTC'
        local_tz = pytz.timezone(user_tz)

        # Convertir a timezone local
        if original_datetime.tzinfo is None:
            local_dt = local_tz.localize(original_datetime)
        else:
            local_dt = original_datetime.astimezone(local_tz)

        target_date = local_dt.date()
        _logger.info("    Fecha objetivo: %s", target_date)

        # Obtener horario de trabajo
        schedule = self._get_work_schedule_for_date(employee, target_date)
        if not schedule:
            _logger.warning("    No se encontró horario para empleado %s en fecha %s",
                            employee.name, target_date)
            return original_datetime

        _logger.info("    Horarios encontrados: %s", schedule)

        calendar = employee.resource_calendar_id

        # Para turnos nocturnos, manejar horarios que cruzan medianoche
        if calendar.nocturna:
            _logger.info("    Turno nocturno detectado - Manejo especial de horarios")
            # Para turnos nocturnos, buscar el horario más apropiado según la hora de la marca
            if local_dt.hour < 12:  # Marca de madrugada
                _logger.info("    Marca de madrugada (%s), buscando horario nocturno",
                             local_dt.hour)
                # Buscar horario del día anterior que termine en la madrugada
                prev_date = target_date - timedelta(days=1)
                prev_schedule = self._get_work_schedule_for_date(employee, prev_date)
                if prev_schedule:
                    _logger.info("    Horarios del día anterior: %s", prev_schedule)
                    # Para madrugada, usar horarios del día anterior
                    schedule = prev_schedule
            else:  # Marca de tarde/noche
                _logger.info("    Marca de tarde/noche (%s), usando horarios del día",
                             local_dt.hour)
                # Para tarde/noche, usar horarios del día actual
        else:
            # Para turnos diurnos, solo usar el primer bloque para la entrada
            if len(schedule) > 1:
                schedule = sorted(schedule, key=lambda x: x['hour_from'])
                first_block = schedule[0]
                _logger.info("    Turno diurno: Usando solo primer bloque: %s",
                             first_block)
                schedule = [first_block]

        # Encontrar el horario de entrada más cercano
        best_match = None
        min_diff = float('inf')

        _logger.info("    Evaluando %d horarios...", len(schedule))

        for i, sched in enumerate(schedule):
            # Convertir hora del horario a datetime
            hour_from_decimal = sched['hour_from']
            hours = int(hour_from_decimal)
            minutes = int((hour_from_decimal - hours) * 60)

            _logger.info("    Horario %d: %.2f -> %02d:%02d", i + 1, hour_from_decimal,
                         hours, minutes)

            # Crear datetime para la hora de inicio del turno
            schedule_start = local_dt.replace(hour=hours, minute=minutes, second=0,
                                              microsecond=0)

            # Para turnos nocturnos que inician el día anterior
            if calendar.nocturna and hours >= 16:  # Turno que inicia desde las 16:00
                if local_dt.hour < 12:  # Si la marca es en la madrugada
                    schedule_start = schedule_start - timedelta(days=1)
                    _logger.info("    Ajustado para turno nocturno: %s", schedule_start)

            # Calcular diferencia
            diff = abs((local_dt - schedule_start).total_seconds())
            _logger.info("    Diferencia con horario %d: %.2f segundos", i + 1, diff)

            if diff < min_diff:
                min_diff = diff
                best_match = schedule_start
                _logger.info("    ✓ Nuevo mejor match: %s (diff: %.2f)", best_match,
                             min_diff)

        if best_match:
            # REGLA CLAVE: Solo ajustar si llegó ANTES del horario
            if local_dt < best_match:
                _logger.info(
                    "    ✅ Llegó ANTES del horario (%s < %s) - AJUSTANDO a horario",
                    local_dt, best_match)
                return best_match.replace(tzinfo=None)
            else:
                _logger.info(
                    "    ❌ Llegó DESPUÉS del horario (%s >= %s) - MANTENIENDO hora real",
                    local_dt, best_match)
                return original_datetime

        _logger.warning("    No se encontró horario válido, manteniendo hora original")
        return original_datetime

    def action_import(self):
        self.ensure_one()

        # Contadores para reporte
        imported_count = 0
        skipped_count = 0
        error_count = 0

        # Arreglos para capturar detalles de errores
        error_details = []
        skipped_details = []

        try:
            # 1) Abrir el Excel
            data = base64.b64decode(self.file_data)
            wb = load_workbook(filename=io.BytesIO(data), data_only=True)
            sheet = wb.active

            # 2) Parámetros de columnas
            IDX_TIEMPO = 0  # Columna A
            IDX_ID = 6  # Columna G (ID/barcode)

            # 3) Timezone del usuario
            user_tz = self.env.user.tz or self.env.context.get('tz') or 'UTC'
            local_tz = pytz.timezone(user_tz)
            utc_tz = pytz.utc

            # 4) Recolectar todos los timestamps por empleado (barcode)
            attend_list = {}
            row_count = 0
            empty_rows = 0

            _logger.info("=== INICIANDO LECTURA DEL EXCEL ===")

            for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if idx == 1:
                    _logger.info("Encabezados detectados: %s", row)
                    continue  # salto encabezado

                row_count += 1
                raw = row[IDX_TIEMPO] if len(row) > IDX_TIEMPO else None
                emp_id = row[IDX_ID] if len(row) > IDX_ID else None

                _logger.info("Fila %d: Tiempo=%s, ID=%s", idx, raw, emp_id)

                if not raw or not emp_id:
                    empty_rows += 1
                    _logger.warning("Fila %d omitida: Tiempo=%s, ID=%s", idx, raw, emp_id)
                    continue

                # parseo a datetime
                try:
                    if isinstance(raw, str):
                        try:
                            tiempo = datetime.fromisoformat(raw)
                            _logger.info("Tiempo parseado desde string: %s -> %s", raw,
                                         tiempo)
                        except ValueError:
                            tiempo = fields.Datetime.from_string(raw)
                            _logger.info("Tiempo parseado con Odoo: %s -> %s", raw,
                                         tiempo)
                    else:
                        tiempo = raw  # openpyxl ya lo entrega como datetime
                        _logger.info("Tiempo ya es datetime: %s", tiempo)
                except Exception as e:
                    _logger.error("Error parseando tiempo en fila %d: %s - Error: %s",
                                  idx, raw, str(e))
                    error_count += 1
                    error_details.append({
                        'tipo': 'Error parseando fecha',
                        'fila': idx,
                        'empleado': emp_id,
                        'tiempo': raw,
                        'error': str(e)
                    })
                    continue

                barcode = str(emp_id).strip()
                attend_list.setdefault(barcode, []).append(tiempo)
                _logger.info("Agregado: Barcode=%s, Tiempo=%s", barcode, tiempo)

            _logger.info("=== RESUMEN LECTURA ===")
            _logger.info("Filas procesadas: %d", row_count)
            _logger.info("Filas vacías omitidas: %d", empty_rows)
            _logger.info("Empleados con marcas: %d", len(attend_list))
            for barcode, times in attend_list.items():
                _logger.info("Empleado %s: %d marcas", barcode, len(times))

            # 5) Para cada empleado, procesar marcas inteligentemente
            _logger.info("=== INICIANDO PROCESAMIENTO POR EMPLEADO ===")

            for barcode, times in attend_list.items():
                try:
                    _logger.info("\n--- Procesando empleado: %s ---", barcode)
                    _logger.info("Marcas originales: %s", times)

                    # 5.1) ordenar
                    times_sorted = sorted(times)
                    _logger.info("Marcas ordenadas: %s", times_sorted)

                    # 5.2) filtrar duplicados dentro de 10 segundos
                    umbral = timedelta(seconds=10)
                    filtered = []
                    for t in times_sorted:
                        if not filtered or (t - filtered[-1]) > umbral:
                            filtered.append(t)
                        else:
                            _logger.info("Marca duplicada omitida: %s (muy cerca de %s)",
                                         t, filtered[-1])
                    times_sorted = filtered
                    _logger.info("Marcas después de filtrar duplicados: %s", times_sorted)

                    emp = self.env["hr.employee"].search([("barcode", "=", barcode)],
                                                         limit=1)
                    if not emp:
                        _logger.error("ERROR: Empleado no encontrado para barcode '%s'",
                                      barcode)
                        error_count += 1
                        error_details.append({
                            'tipo': 'Empleado no encontrado',
                            'empleado': barcode,
                            'marcas': times,
                            'error': f'No existe empleado con barcode {barcode}'
                        })
                        continue

                    _logger.info("Empleado encontrado: %s (ID: %d)", emp.name, emp.id)

                    # PASO 1: Analizar horario del empleado
                    schedule_info = self._analyze_employee_schedule(emp)
                    if not schedule_info:
                        _logger.error(
                            "ERROR: Empleado %s (barcode: %s) no tiene horario de trabajo asignado",
                            emp.name, barcode)
                        error_count += 1
                        error_details.append({
                            'tipo': 'Sin horario de trabajo',
                            'empleado': f'{emp.name} ({barcode})',
                            'marcas': times,
                            'error': 'Empleado no tiene horario de trabajo asignado'
                        })
                        continue

                    calendar = schedule_info['calendar']
                    _logger.info("Horario de trabajo: %s (Nocturno: %s)", calendar.name,
                                 calendar.nocturna)

                    # PASO 2: Agrupar marcas por turno de trabajo (considerando turnos nocturnos)
                    daily_marks = self._group_marks_by_work_shift(times_sorted,
                                                                  schedule_info)
                    _logger.info("Marcas agrupadas por turno de trabajo: %s", daily_marks)

                    # PASO 3: Procesar cada turno con la información del horario
                    for work_date, day_times in daily_marks.items():
                        _logger.info("\n  -- Procesando turno: %s --", work_date)
                        _logger.info("  Marcas del turno: %s", day_times)

                        # Para turnos nocturnos, ordenar las marcas cronológicamente
                        # considerando que pueden abarcar dos días
                        if calendar.nocturna and len(day_times) > 1:
                            # Separar marcas por tipo para ordenar correctamente
                            evening_marks = []  # Marcas de tarde/noche del primer día
                            morning_marks = []  # Marcas de madrugada del segundo día

                            for mark in day_times:
                                if mark.hour >= 16:  # Tarde/noche (desde 16:00)
                                    evening_marks.append(mark)
                                elif mark.hour <= 10:  # Madrugada
                                    morning_marks.append(mark)
                                else:  # Día (caso raro)
                                    morning_marks.append(mark)

                            # Ordenar cada grupo por separado y luego concatenar
                            evening_marks.sort()
                            morning_marks.sort()
                            day_times = evening_marks + morning_marks

                            _logger.info("  Marcas reordenadas para turno nocturno:")
                            _logger.info("    Tarde/noche: %s", evening_marks)
                            _logger.info("    Madrugada: %s", morning_marks)
                            _logger.info("    Orden final: %s", day_times)

                        # Verificar duplicados completos para este día
                        sample_time = day_times[0]
                        if self._check_duplicate_attendance(emp, sample_time):
                            _logger.info(
                                "  DUPLICADO: Asistencia completa ya existe para %s en %s - Omitiendo",
                                emp.name, work_date)
                            skipped_count += 1
                            skipped_details.append({
                                'empleado': f'{emp.name} ({barcode})',
                                'fecha': work_date,
                                'marcas': day_times,
                                'motivo': 'Asistencia completa ya existe'
                            })
                            continue

                        # Buscar asistencias parciales existentes para completar
                        existing_partial = self._find_existing_partial_attendance(emp,
                                                                                  work_date)

                        # PASO 4: Clasificar marcas según horario
                        classified_marks = self._classify_marks_by_schedule(day_times,
                                                                            schedule_info)

                        _logger.info("  Marcas clasificadas: %s", {
                            'entradas': len(classified_marks['entries']),
                            'salidas': len(classified_marks['exits']),
                            'desconocidas': len(classified_marks['unknown'])
                        })

                        # PASO 5: Determinar qué crear/completar
                        if existing_partial:
                            # Completar asistencia parcial existente
                            if classified_marks['exits']:
                                exit_time = classified_marks['exits'][-1]  # Última salida
                                _logger.info(
                                    "  Completando asistencia parcial con salida: %s",
                                    exit_time)

                                try:
                                    dt_out_utc = local_tz.localize(exit_time).astimezone(
                                        utc_tz) if exit_time.tzinfo is None else exit_time.astimezone(
                                        utc_tz)
                                    existing_partial.write({
                                        'check_out': dt_out_utc.strftime(
                                            "%Y-%m-%d %H:%M:%S"),
                                        'is_partial': False,
                                        'partial_type': 'complete'
                                    })
                                    _logger.info(
                                        "  ✓ Asistencia parcial completada (ID: %d)",
                                        existing_partial.id)
                                    imported_count += 1
                                    continue
                                except Exception as e:
                                    _logger.error(
                                        "  ERROR completando asistencia parcial: %s",
                                        str(e))
                                    error_count += 1
                                    error_details.append({
                                        'tipo': 'Error completando asistencia parcial',
                                        'empleado': f'{emp.name} ({barcode})',
                                        'fecha': work_date,
                                        'salida': exit_time,
                                        'error': str(e)
                                    })
                                    continue
                            else:
                                _logger.info(
                                    "  Asistencia parcial existente, pero no hay marcas de salida nuevas")
                                skipped_count += 1
                                skipped_details.append({
                                    'empleado': f'{emp.name} ({barcode})',
                                    'fecha': work_date,
                                    'marcas': day_times,
                                    'motivo': 'Asistencia parcial existente sin nuevas salidas'
                                })
                                continue

                        # PASO 6: Crear nueva asistencia
                        entries = classified_marks['entries']
                        exits = classified_marks['exits']

                        if entries and exits:
                            # Asistencia completa
                            entry_time = entries[0]  # Primera entrada
                            exit_time = exits[-1]  # Última salida

                            _logger.info("  Creando asistencia completa: %s -> %s",
                                         entry_time, exit_time)

                            # Para turnos nocturnos, ajustar fechas si es necesario
                            if calendar.nocturna:
                                # Si la entrada es de tarde (>=16h) y salida es de madrugada (<=10h)
                                # la salida es del día siguiente
                                if entry_time.hour >= 16 and exit_time.hour <= 10:
                                    # Verificar si ambas marcas tienen la misma fecha (incorrecto para nocturno)
                                    if entry_time.date() == exit_time.date():
                                        # Corregir: la salida debería ser del día siguiente
                                        _logger.info(
                                            "  Turno nocturno: Ajustando salida al día siguiente")
                                        exit_time = exit_time + timedelta(days=1)
                                        _logger.info("  Salida ajustada: %s", exit_time)
                                    # Si ya tienen fechas diferentes, verificar que sea correcto
                                    elif (exit_time.date() - entry_time.date()).days != 1:
                                        _logger.warning(
                                            "  Turno nocturno: Diferencia de fechas inusual entre entrada (%s) y salida (%s)",
                                            entry_time.date(), exit_time.date())

                            # Procesar entrada y salida
                            adjusted_check_in = self._adjust_check_in_time(emp,
                                                                           entry_time)

                            try:
                                # Convertir a UTC
                                dt_in_utc = local_tz.localize(
                                    adjusted_check_in).astimezone(
                                    utc_tz) if adjusted_check_in.tzinfo is None else adjusted_check_in.astimezone(
                                    utc_tz)
                                dt_out_utc = local_tz.localize(exit_time).astimezone(
                                    utc_tz) if exit_time.tzinfo is None else exit_time.astimezone(
                                    utc_tz)
                                real_check_in_utc = local_tz.localize(
                                    entry_time).astimezone(
                                    utc_tz) if entry_time.tzinfo is None else entry_time.astimezone(
                                    utc_tz)

                                attendance_vals = {
                                    "employee_id": emp.id,
                                    "check_in": real_check_in_utc.strftime(
                                        "%Y-%m-%d %H:%M:%S"),
                                    "check_out": dt_out_utc.strftime("%Y-%m-%d %H:%M:%S"),
                                    "check_in_schedule": dt_in_utc.strftime(
                                        "%Y-%m-%d %H:%M:%S"),
                                    "is_partial": False,
                                    "partial_type": "complete"
                                }

                                _logger.info("  Valores finales: IN=%s, OUT=%s, SCH=%s",
                                             real_check_in_utc, dt_out_utc, dt_in_utc)

                                new_attendance = self.env["hr.attendance"].create(
                                    attendance_vals)
                                _logger.info("  ✓ Asistencia completa creada (ID: %d)",
                                             new_attendance.id)
                                imported_count += 1

                            except Exception as e:
                                _logger.error("  ERROR creando asistencia completa: %s",
                                              str(e))
                                error_count += 1
                                error_details.append({
                                    'tipo': 'Error creando asistencia completa',
                                    'empleado': f'{emp.name} ({barcode})',
                                    'fecha': work_date,
                                    'entrada': entry_time,
                                    'salida': exit_time,
                                    'error': str(e)
                                })
                                continue

                        elif entries:
                            # Solo entrada - crear asistencia parcial
                            entry_time = entries[0]
                            _logger.info(
                                "  Creando asistencia parcial (solo entrada): %s",
                                entry_time)

                            # Activar automáticamente allow_partial_attendance si no está activo
                            if not self.allow_partial_attendance:
                                _logger.info(
                                    "  Activando automáticamente 'allow_partial_attendance' debido a marcas parciales detectadas")
                                self.allow_partial_attendance = True

                            # Ajustar la entrada según el horario programado
                            adjusted_check_in = self._adjust_check_in_time(emp,
                                                                           entry_time)

                            try:
                                # Convertir a UTC
                                dt_in_utc = local_tz.localize(
                                    adjusted_check_in).astimezone(
                                    utc_tz) if adjusted_check_in.tzinfo is None else adjusted_check_in.astimezone(
                                    utc_tz)
                                real_check_in_utc = local_tz.localize(
                                    entry_time).astimezone(
                                    utc_tz) if entry_time.tzinfo is None else entry_time.astimezone(
                                    utc_tz)

                                attendance_vals = {
                                    "employee_id": emp.id,
                                    "check_in": real_check_in_utc.strftime(
                                        "%Y-%m-%d %H:%M:%S"),
                                    "check_in_schedule": dt_in_utc.strftime(
                                        "%Y-%m-%d %H:%M:%S"),
                                    "is_partial": True,
                                    "partial_type": "entry_only"
                                    # check_out se deja en blanco intencionalmente
                                }

                                new_attendance = self.env["hr.attendance"].create(
                                    attendance_vals)
                                _logger.info(
                                    "  ✓ Asistencia parcial creada (ID: %d) - Entrada ajustada a horario, salida en blanco",
                                    new_attendance.id)
                                imported_count += 1

                            except Exception as e:
                                _logger.error("  ERROR creando asistencia parcial: %s",
                                              str(e))
                                error_count += 1
                                error_details.append({
                                    'tipo': 'Error creando asistencia parcial',
                                    'empleado': f'{emp.name} ({barcode})',
                                    'fecha': work_date,
                                    'entrada': entry_time,
                                    'error': str(e)
                                })
                                continue

                        elif exits:
                            # Solo salida - crear asistencia parcial con entrada según horario
                            exit_time = exits[0]
                            _logger.info(
                                "  Creando asistencia parcial (solo salida): %s",
                                exit_time)

                            # Activar automáticamente allow_partial_attendance si no está activo
                            if not self.allow_partial_attendance:
                                _logger.info(
                                    "  Activando automáticamente 'allow_partial_attendance' debido a marcas parciales detectadas")
                                self.allow_partial_attendance = True

                            # Calcular la entrada teórica según el horario de trabajo
                            theoretical_entry = self._calculate_theoretical_entry_time(
                                emp, work_date, exit_time, schedule_info)

                            if theoretical_entry:
                                try:
                                    # Convertir a UTC
                                    dt_in_utc = local_tz.localize(
                                        theoretical_entry).astimezone(
                                        utc_tz) if theoretical_entry.tzinfo is None else theoretical_entry.astimezone(
                                        utc_tz)
                                    dt_out_utc = local_tz.localize(exit_time).astimezone(
                                        utc_tz) if exit_time.tzinfo is None else exit_time.astimezone(
                                        utc_tz)

                                    attendance_vals = {
                                        "employee_id": emp.id,
                                        "check_in": dt_in_utc.strftime(
                                            "%Y-%m-%d %H:%M:%S"),
                                        "check_in_schedule": dt_in_utc.strftime(
                                            "%Y-%m-%d %H:%M:%S"),
                                        "check_out": dt_out_utc.strftime(
                                            "%Y-%m-%d %H:%M:%S"),
                                        "is_partial": True,
                                        "partial_type": "exit_only"
                                    }

                                    new_attendance = self.env["hr.attendance"].create(
                                        attendance_vals)
                                    _logger.info(
                                        "  ✓ Asistencia parcial creada (ID: %d) - Entrada teórica según horario, salida real",
                                        new_attendance.id)
                                    imported_count += 1

                                except Exception as e:
                                    _logger.error(
                                        "  ERROR creando asistencia parcial (solo salida): %s",
                                        str(e))
                                    error_count += 1
                                    error_details.append({
                                        'tipo': 'Error creando asistencia parcial (solo salida)',
                                        'empleado': f'{emp.name} ({barcode})',
                                        'fecha': work_date,
                                        'salida': exit_time,
                                        'error': str(e)
                                    })
                                    continue
                            else:
                                # No se pudo calcular entrada teórica
                                _logger.warning(
                                    "  No se pudo calcular entrada teórica para empleado %s en fecha %s",
                                    emp.name, work_date)
                                error_count += 1
                                error_details.append({
                                    'tipo': 'No se pudo calcular entrada teórica',
                                    'empleado': f'{emp.name} ({barcode})',
                                    'fecha': work_date,
                                    'salida': exit_time,
                                    'error': 'No se encontró horario válido para calcular entrada teórica'
                                })
                                continue

                        else:
                            # No hay entradas ni salidas válidas - esto es un error
                            error_type = 'Sin marcas válidas'
                            error_msg = 'No se encontraron marcas de entrada ni salida válidas según el horario'

                            _logger.warning(
                                "  No se pueden procesar marcas para %s en %s: %s",
                                emp.name, work_date, error_msg)

                            error_count += 1
                            error_details.append({
                                'tipo': error_type,
                                'empleado': f'{emp.name} ({barcode})',
                                'fecha': work_date,
                                'marcas': day_times,
                                'clasificacion': classified_marks,
                                'error': error_msg
                            })
                            continue

                except Exception as e:
                    _logger.error("ERROR general procesando empleado %s: %s", barcode,
                                  str(e))
                    error_count += 1
                    error_details.append({
                        'tipo': 'Error general procesando empleado',
                        'empleado': barcode,
                        'marcas': times if 'times' in locals() else 'No disponibles',
                        'error': str(e)
                    })
                    continue

        except Exception as e:
            _logger.error("Error general en importación: %s", str(e))
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Error',
                    'type': 'danger',
                }
            }

        # 6) Mostrar resumen de importación con detalles
        message = f"Importación completada:\n"
        message += f"• Registros importados: {imported_count}\n"
        message += f"• Registros omitidos (duplicados): {skipped_count}\n"
        # message += f"• Errores: {error_count}\n\n"

        # # Agregar detalles de errores
        # if error_details:
        #     message += "=== DETALLE DE ERRORES ===\n"
        #     error_summary = {}
        #     for error in error_details:
        #         error_type = error['tipo']
        #         # Filtrar errores de asistencias parciales
        #         if error_type in ['Solo entrada, asistencias parciales deshabilitadas',
        #                           'Solo entrada, falta salida',
        #                           'Error creando asistencia parcial']:
        #             continue
        #
        #         if error_type not in error_summary:
        #             error_summary[error_type] = []
        #         error_summary[error_type].append(error)
        #
        #     for error_type, errors in error_summary.items():
        #         message += f"\n{error_type} ({len(errors)} casos):\n"
        #         for i, error in enumerate(errors[:10]):  # Mostrar solo los primeros 10
        #             if 'empleado' in error:
        #                 # Solo mostrar el nombre del empleado
        #                 employee_name = error['empleado']
        #                 # Si es un barcode (solo números), mostrar tal como está
        #                 if employee_name.isdigit():
        #                     message += f"  - Barcode: {employee_name}\n"
        #                 else:
        #                     # Extraer solo el nombre, sin el barcode entre paréntesis
        #                     if '(' in employee_name and ')' in employee_name:
        #                         name_only = employee_name.split('(')[0].strip()
        #                         message += f"  - {name_only}\n"
        #                     else:
        #                         message += f"  - {employee_name}\n"
        #             else:
        #                 message += f"  - Fila {error.get('fila', 'N/A')}\n"
        #         if len(errors) > 10:
        #             message += f"  ... y {len(errors) - 10} casos más\n"
        #
        # # Agregar detalles de omitidos (solo los que no son por asistencias parciales)
        # if skipped_details:
        #     filtered_skipped = [s for s in skipped_details if
        #                         'asistencia parcial' not in s['motivo'].lower()]
        #     if filtered_skipped:
        #         message += f"\n=== DETALLE DE OMITIDOS ===\n"
        #         for i, skipped in enumerate(
        #                 filtered_skipped[:10]):  # Mostrar solo los primeros 10
        #             employee_name = skipped['empleado']
        #             # Extraer solo el nombre, sin el barcode
        #             if '(' in employee_name and ')' in employee_name:
        #                 name_only = employee_name.split('(')[0].strip()
        #                 message += f"  - {name_only}\n"
        #             else:
        #                 message += f"  - {employee_name}\n"
        #         if len(filtered_skipped) > 10:
        #             message += f"  ... y {len(filtered_skipped) - 10} casos más\n"

        # Log completo para debugging
        _logger.info("=== RESUMEN COMPLETO DE ERRORES ===")
        _logger.info("Errores detallados: %s", error_details)
        _logger.info("Omitidos detallados: %s", skipped_details)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Importación Completada',
                'message': message,
                'type': 'success' if error_count == 0 else 'warning',
                'sticky': True,
            }
        }