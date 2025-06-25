# models/attendance_report.py
from odoo import models, fields, api
from datetime import datetime, timedelta
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pytz


class AttendanceReportWizard(models.TransientModel):
    _name = 'attendance.report.wizard'
    _description = 'Reporte de Asistencias por Rango de Fechas'

    date_from = fields.Date(
        string='Fecha Desde',
        required=True,
        default=lambda self: fields.Date.today().replace(day=1)
    )
    date_to = fields.Date(
        string='Fecha Hasta',
        required=True,
        default=fields.Date.today
    )
    employee_ids = fields.Many2many(
        'hr.employee',
        string='Empleados',
        help='Dejar vacío para incluir todos los empleados'
    )
    department_ids = fields.Many2many(
        'hr.department',
        string='Departamentos',
        help='Dejar vacío para incluir todos los departamentos'
    )

    @api.onchange('department_ids')
    def _onchange_department_ids(self):
        """Filtrar empleados por departamentos seleccionados"""
        if self.department_ids:
            return {
                'domain': {
                    'employee_ids': [('department_id', 'in', self.department_ids.ids)]
                }
            }
        else:
            return {
                'domain': {
                    'employee_ids': []
                }
            }

    def action_generate_report(self):
        """Genera el reporte de asistencias en Excel"""
        return self._generate_excel_report()

    def _generate_excel_report(self):
        """Genera el reporte en Excel usando openpyxl"""
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Asistencias"

        # Definir estilos
        title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4',
                                 fill_type='solid')
        title_alignment = Alignment(horizontal='center', vertical='center')

        header_font = Font(name='Arial', size=12, bold=True)
        header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3',
                                  fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        employee_font = Font(name='Arial', size=11, bold=True)
        employee_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2',
                                    fill_type='solid')

        data_alignment = Alignment(horizontal='center')
        number_alignment = Alignment(horizontal='right')

        total_font = Font(name='Arial', size=10, bold=True)
        total_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C',
                                 fill_type='solid')

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Configurar anchos de columna (ajustados para fecha y hora completa)
        column_widths = [20, 20, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Título principal
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = 'REPORTE DE ASISTENCIAS'
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_alignment

        ws.merge_cells('A2:C2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = f'Del {self.date_from.strftime("%d/%m/%Y")} al {self.date_to.strftime("%d/%m/%Y")}'
        subtitle_cell.font = title_font
        subtitle_cell.fill = title_fill
        subtitle_cell.alignment = title_alignment

        # Obtener datos
        attendance_data = self._get_attendance_data()

        current_row = 4

        for employee_data in attendance_data:
            # Nombre del empleado con departamento
            department_name = employee_data["employee"].department_id.name if \
            employee_data["employee"].department_id else "Sin Departamento"
            ws.merge_cells(f'A{current_row}:C{current_row}')
            emp_cell = ws[f'A{current_row}']
            emp_cell.value = f'EMPLEADO: {employee_data["employee"].name} - {employee_data["employee"].identification_id or "Sin ID"} | DEPTO: {department_name}'
            emp_cell.font = employee_font
            emp_cell.fill = employee_fill
            emp_cell.border = border
            current_row += 1

            # Encabezados (solo 3 columnas)
            headers = ['Entrada', 'Salida', 'Horas Trabajadas']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            current_row += 1

            # Datos de asistencia
            for attendance in employee_data['attendances']:
                # Entrada (fecha y hora completa en zona horaria del usuario)
                entrada_str = attendance['check_in_local'].strftime('%d/%m/%Y %H:%M') if \
                attendance['check_in_local'] else ''
                ws.cell(row=current_row, column=1).value = entrada_str
                ws.cell(row=current_row, column=1).alignment = data_alignment
                ws.cell(row=current_row, column=1).border = border

                # Salida (fecha y hora completa en zona horaria del usuario)
                salida_str = attendance['check_out_local'].strftime('%d/%m/%Y %H:%M') if \
                attendance['check_out_local'] else 'Sin salida'
                ws.cell(row=current_row, column=2).value = salida_str
                ws.cell(row=current_row, column=2).alignment = data_alignment
                ws.cell(row=current_row, column=2).border = border

                # Horas trabajadas
                ws.cell(row=current_row, column=3).value = attendance['worked_hours']
                ws.cell(row=current_row, column=3).alignment = number_alignment
                ws.cell(row=current_row, column=3).number_format = '0.00'
                ws.cell(row=current_row, column=3).border = border

                current_row += 1

            # Fila de totales (solo 3 columnas)
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                cell.font = total_font
                cell.fill = total_fill
                cell.border = border

                if col == 1:
                    cell.value = 'TOTALES:'
                elif col == 2:
                    cell.value = ''
                elif col == 3:
                    cell.value = employee_data['total_worked_hours']
                    cell.alignment = number_alignment
                    cell.number_format = '0.00'

            current_row += 2

            current_row += 1

        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Crear archivo adjunto
        filename = f'reporte_asistencias_{self.date_from.strftime("%Y%m%d")}_{self.date_to.strftime("%Y%m%d")}.xlsx'
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }

    def _get_attendance_data(self):
        """Obtiene y procesa los datos de asistencia"""
        domain = [
            ('check_in', '>=', self.date_from),
            ('check_in', '<=', self.date_to)
        ]

        if self.employee_ids:
            domain.append(('employee_id', 'in', self.employee_ids.ids))

        if self.department_ids:
            domain.append(('employee_id.department_id', 'in', self.department_ids.ids))

        attendances = self.env['hr.attendance'].search(domain,
                                                       order='employee_id, check_in')

        # Obtener zona horaria del usuario
        user_tz = pytz.timezone(self.env.user.tz or 'UTC')

        # Agrupar por empleado
        employee_data = {}
        for attendance in attendances:
            emp_id = attendance.employee_id.id
            if emp_id not in employee_data:
                employee_data[emp_id] = {
                    'employee': attendance.employee_id,
                    'attendances': [],
                    'total_worked_hours': 0.0,
                }

            # Convertir fechas UTC a zona horaria local
            check_in_local = None
            check_out_local = None

            if attendance.check_in:
                check_in_utc = attendance.check_in.replace(tzinfo=pytz.UTC)
                check_in_local = check_in_utc.astimezone(user_tz)

            if attendance.check_out:
                check_out_utc = attendance.check_out.replace(tzinfo=pytz.UTC)
                check_out_local = check_out_utc.astimezone(user_tz)

            # Calcular horas trabajadas usando las fechas UTC originales
            worked_hours = 0.0
            if attendance.check_in and attendance.check_out:
                delta = attendance.check_out - attendance.check_in
                worked_hours = delta.total_seconds() / 3600.0

            attendance_data = {
                'employee_id': attendance.employee_id.name,
                'check_in': attendance.check_in,  # UTC original para cálculos
                'check_out': attendance.check_out,  # UTC original para cálculos
                'check_in_local': check_in_local,  # Local para mostrar
                'check_out_local': check_out_local,  # Local para mostrar
                'worked_hours': worked_hours,
            }

            employee_data[emp_id]['attendances'].append(attendance_data)
            employee_data[emp_id]['total_worked_hours'] += worked_hours

        # Aplicar límite de 70 horas
        for emp_data in employee_data.values():
            if emp_data['total_worked_hours'] > 70:
                emp_data['total_worked_hours'] = 70.0

        return list(employee_data.values())